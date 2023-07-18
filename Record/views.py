from django.shortcuts import render,redirect
from django.contrib.auth.decorators import login_required
from django.http import HttpResponseBadRequest
import xlrd
from openpyxl import load_workbook
import datetime

from Batch.models import Batch,DeductionCode
from Company.models import Company
from Record.models import Record
from Record.forms import RecordForm,UploadForm


@login_required
def import_data(request):
    if request.method == 'POST':
        # Get the uploaded file
        file = request.FILES['batch_file']
        code = request.POST.get('select')
        # Open the file using openpyxl
        workbook = load_workbook(filename=file, read_only=True)
        worksheet = workbook.active

        #get company instance
        company_id = request.session.get('company_id')
        company = Company.objects.get(id=company_id) 
        # Get the deduction code and create a new batch
        deduction_code = DeductionCode.objects.get(id=int(code))
        new_batch = Batch.objects.create(deduction_code=deduction_code,company=company)

        # Loop through the rows and save each record to the database
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            reference = row[0]
            id_number = row[1]
            ec_number = row[2]
            type = row[3]
            start_date = row[4]
            end_date = row[5]
            # start_date = datetime.datetime.strptime(str(row[4]), '%d-%m-%Y').strftime('%Y-%m-%d')
            # end_date = datetime.datetime.strptime(str(row[5]), '%d-%m-%Y').strftime('%Y-%m-%d')
            amount = row[6]
            if not (reference == None or type == None or ec_number == None):
                Record.objects.create(
                    batch=new_batch,
                    deduction_code=deduction_code,
                    code=deduction_code.code,
                    request_type=type,
                    ec_number=ec_number,
                    id_number=id_number,
                    transaction_refence=reference,
                    deductions_start_date=start_date,
                    deductions_end_date=end_date,
                    installment_amount=amount
                )
        # Redirect to a success page
        return redirect('record:records')

    # Render the form if the request method is GET
    codes = DeductionCode.objects.all()
    return render(request, 'dashboard/upload-batch.html', {'codes': codes})

@login_required
def record_registration(request):
    if request.method =='POST':
        form = RecordForm(request.POST)
        if form.is_valid():
            record = form.save(commit=False)
            #get company instance
            company_id = request.session.get('company_id')
            company = Company.objects.get(id=company_id) 
            batch = Batch.objects.filter(company=company,status = Batch.Status.DRAFT).last()
            if batch:
                record.batch = batch
                record.record_source = "Individual Upload"
                record.save()
            else:
                dd_code =request.POST.get('deduction_code')
                deduction_code = DeductionCode.objects.get(id=dd_code)
                new_batch = Batch.objects.create(deduction_code = deduction_code,company=company)
                record.batch = new_batch
                record.record_source = "Individual Upload"
                record.save()
            return redirect("record:records")
    else:
        form = RecordForm()
    return render(request,'dashboard/add-record.html',{'form':form})

@login_required
def get_records(request):
    company_id = request.session.get('company_id')
    company = Company.objects.get(id=company_id) 
    records = Record.objects.filter(batch__company = company).select_related('batch')
    return render(request,'dashboard/records.html',{'records':records,'company':company})

@login_required
def filter_type(request):
    selected_option = request.GET.get('type')
    # Perform the necessary query based on the selected option
    if selected_option == 'NEW':
        records = Record.objects.filter(request_type='NEW')
    elif selected_option == 'CHANGE':
        records= Record.objects.filter(request_type='CHANGE')
    elif selected_option == 'DELETE':
        records = Record.objects.filter(status='DELETE')
    return render(request,'dashboard/records.html',{'records':records})

@login_required
def filter_status(request):
    selected_option = request.GET.get('type')
    # Perform the necessary query based on the selected option
    if selected_option == 'CANCELED':
        records = Record.objects.filter(status='CANCELED')
    elif selected_option == 'DRAFT':
        records= Record.objects.filter(status='DRAFT')
    elif selected_option == 'FAILED':
        records = Record.objects.filter(status='FAILED')
    elif selected_option == 'PROCESSED':
        records = Record.objects.filter(status='PROCESSED')
    elif selected_option == 'PROCESSING':
        records = Record.objects.filter(status='PROCESSING')
    elif selected_option == 'SUCCESS':
        records = Record.objects.filter(status='SUCCESS')
    elif selected_option == 'SENT':
        records = Record.objects.filter(status='SENT')
    elif selected_option == 'SAVED':
        records = Record.objects.filter(status='SAVED')
        
    return render(request,'dashboard/records.html',{'records':records})
    
@login_required
def record_detail(request, record_id):
    record = Record.objects.get(record_id=record_id)
    return render(request,'dashboard/record.html',{'record':record})