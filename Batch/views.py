from django.shortcuts import render,redirect, get_object_or_404
from django.db.models import Sum

from openpyxl import load_workbook

# Create your views here.

from Batch.models import Batch, DeductionCode
from Batch.forms import DeductionCodeForm
from Company.models import Company
from Response.models import Response

def deduction_code_registration(request):
    if request.method =='POST':
        form = DeductionCodeForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect("batch:codes")
    else:
        form = DeductionCodeForm()
    return render(request,'deductioncode/add-deduction-code.html',{'form':form})

# @login_required
def get_deduction_codes(request):
    codes = DeductionCode.objects.all()
    return render(request,'deductioncode/deduction-codes.html',{'codes':codes})


def all_batches(request):
    company_id = request.session.get('company_id')
    company = Company.objects.get(id=company_id)
    batches = company.company_batches.all()
    return render(request,'dashboard/batches.html',{'batches':batches})


def draft_batches(request):
    batches = Batch.draft.all()
    return render(request,'dashboard/drafts.html',{'batches':batches})

def view_batch(request,id):
    batch = get_object_or_404(Batch,id=id)
    total_price = batch.batch_records.aggregate(Sum('installment_amount'))['installment_amount__sum']
    print(total_price)
    #cost = str(float(total_price)/100.00)
    return render(request,'dashboard/batch.html',{'batch':batch,'total_value':str(total_price/100.00)})

def send_batch(request):
    is_one = False
    company_id = request.session.get('company_id')
    company = Company.objects.get(id=company_id)
    batches = company.company_batches.filter(status = Batch.Status.DRAFT)
    if batches.count() == 1:
        is_one = True
    return render(request,'dashboard/drafts.html',{'batches':batches,'is_one':is_one})


# views.py

def upload_excel(request):
    if request.method == 'POST':
        file = request.FILES['excel_file']
        wb = load_workbook(filename=file)
        ws = wb.active
        for row in ws.iter_rows(min_row=2):
            pass
            # data = [cell.value for cell in row]
            # obj = MyModel(field1=data[0], field2=data[1], field3=data[2])
            # obj.save()
        return render(request, 'admin/upload_excel_success.html')
    else:
        return render(request, 'admin/upload_excel_form.html')
    
    
#@login_required
def upload_responses(request,id):
    if request.method == 'POST':
        # Get the uploaded file
        file = request.FILES['excel_file']
        # Open the file using openpyxl
        workbook = load_workbook(filename=file, read_only=True)
        worksheet = workbook.active
        batch = get_object_or_404(Batch,id=id)

        # Loop through the rows and save each record to the database
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            recid = row[0]
            deductionCode = row[1]
            reference = row[2]
            idNumber = row[3]
            ecNumber = row[4]
            type = row[5]
            status = row[6]
            startDate = row[7]
            endDate = row[8]
            amount = row[9]
            name = row[10]
            bankAccount = row[11]
            message = row[12]
            
            #upload record
            Response.objects.create(
                batch=batch,
                recid=recid,
                deductionCode=deductionCode,
                reference = reference,
                idNumber = idNumber,
                ecNumber=ecNumber,
                type = type,
                status = status,
                startDate = startDate,
                endDate = endDate,
                amount = amount,
                name = name,
                bankAccount = bankAccount,
                message = message
            )
        batch.status = Batch.Status.SUCCESS
        batch.save()
            
        # Redirect to a success page
        return render(request, 'admin/upload_excel_success.html')
    else:
        return render(request, 'admin/upload_excel_form.html')
    

    # Render the form if the request method is GET
    codes = DeductionCode.objects.all()
    return render(request, 'dashboard/upload-batch.html', {'codes': codes})

    